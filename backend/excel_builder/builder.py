"""builder.py — monthly Excel workbook builder for FinanceTracker (§7.8, FR-30).

Builds a financetracker-YYYY-MM.xlsx with two sheets:
  - "Transactions" — one row per MonthRow (Date / Description / Amount / Category).
  - "Summary"      — one row per category total, then a Net row.

Privacy note
------------
This file handles FULL LOCAL DATA (raw descriptions, real amounts).
It is the owner's local/own-Drive archive — NOT the sanitised off-machine payload.
Do NOT confuse this with the §7.5 sanitised tuple (row_index, cleaned_description, amount).
Nothing written here is sent off-machine; *.xlsx is gitignored and must never be committed.

Secrets
-------
OUTPUT_DIR is read from .env via python-dotenv; never hardcoded.
No file or directory is created by a bare ``import backend.excel_builder``.
"""
from __future__ import annotations

import os
from decimal import Decimal
from pathlib import Path

import openpyxl
from dotenv import load_dotenv

from backend.store import MonthRow


def resolve_output_dir(override: str | os.PathLike | None = None) -> Path:
    """Resolve the output directory for .xlsx workbooks.

    Priority: override argument > $OUTPUT_DIR environment variable > './output'.
    Calls load_dotenv() (no-op if already loaded). Does NOT create the directory.
    """
    if override is not None:
        return Path(override)
    load_dotenv()  # no-op if already loaded; safe to call multiple times
    output_dir = os.getenv("OUTPUT_DIR", "./output")
    return Path(output_dir)


def build_workbook(
    year_month: str,
    transactions: list[MonthRow],
    summary: dict,
    *,
    output_dir: str | os.PathLike | None = None,
) -> Path:
    """Build the monthly workbook with openpyxl and return the written local Path.

    No network, no Store/DB coupling — caller passes the rows in.

    Parameters
    ----------
    year_month:
        'YYYY-MM' string used for the filename; NOT re-derived from rows.
    transactions:
        List of MonthRow objects to write to the Transactions sheet.
    summary:
        Dict returned by Store.summary(year_month) — expected shape::

            {
                "year_month": "YYYY-MM" | None,
                "totals": {"Category": "amount_str", ...},
                "net": "amount_str",
                "count": N,
            }

        Money values are str(Decimal), never float.
    output_dir:
        Override the output directory. Defaults to OUTPUT_DIR from .env or './output'.

    Returns
    -------
    Path
        Full path to the written .xlsx file.
    """
    resolved_dir = resolve_output_dir(output_dir)
    # Create the directory at WRITE time only — never on import
    resolved_dir.mkdir(parents=True, exist_ok=True)

    filename = f"financetracker-{year_month}.xlsx"
    out_path = resolved_dir / filename

    # ---- Build workbook ----
    # Set first sheet title via wb.active to avoid a stray empty "Sheet"
    wb = openpyxl.Workbook()
    ws_txn = wb.active
    ws_txn.title = "Transactions"
    ws_summary = wb.create_sheet("Summary")

    # ---- Sheet 1: Transactions ----
    ws_txn.append(["Date", "Description", "Amount", "Category"])
    for row in transactions:
        category = row.category if row.category is not None else "Uncategorised"
        # Decimal → float only at the openpyxl boundary for display.
        # The canonical exact value lives in SQLite as str(Decimal); it is never
        # recomputed from this float.  Negative debits are preserved as negative.
        amount_num = float(row.amount)
        ws_txn.append([row.date, row.description, amount_num, category])
        # Apply 2dp number format to the Amount cell (column 3)
        cell = ws_txn.cell(row=ws_txn.max_row, column=3)
        cell.number_format = "0.00"

    # ---- Sheet 2: Summary ----
    ws_summary.append(["Category", "Total"])
    totals: dict[str, str] = summary.get("totals", {})
    for cat_label, total_str in totals.items():
        # Decimal(str) → float only at openpyxl boundary; "-0.00" parses cleanly
        total_num = float(Decimal(total_str))
        ws_summary.append([cat_label, total_num])
        cell = ws_summary.cell(row=ws_summary.max_row, column=2)
        cell.number_format = "0.00"

    # Net row (always present — store guarantees net, default "0.00")
    net_str = summary.get("net", "0.00")
    net_num = float(Decimal(net_str))
    ws_summary.append(["Net", net_num])
    net_cell = ws_summary.cell(row=ws_summary.max_row, column=2)
    net_cell.number_format = "0.00"

    wb.save(str(out_path))
    return out_path
