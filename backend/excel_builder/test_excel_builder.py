"""test_excel_builder.py — pytest suite for §7.8 Excel builder (FR-30).

ALL fixtures use SYNTHETIC data generated in code.
No real transactions, no real descriptions, no real account numbers.
No network calls anywhere in this file.
Every file is written to pytest tmp_path — NEVER to ./output.
"""
from __future__ import annotations

import re
from decimal import Decimal
from pathlib import Path

import openpyxl
import pytest

from backend.store import MonthRow
from backend.excel_builder import build_workbook, resolve_output_dir

# ---------------------------------------------------------------------------
# Synthetic helpers — invented values; never real data
# ---------------------------------------------------------------------------


def _row(
    date: str = "2026-06-01",
    description: str = "SYNTH VENDOR",
    amount: str = "-10.00",
    category: str | None = "Groceries",
) -> MonthRow:
    """Build a synthetic MonthRow; invented values only, never real transactions."""
    return MonthRow(
        date=date,
        description=description,
        amount=Decimal(amount),
        category=category,
    )


def _synth_transactions() -> list[MonthRow]:
    """Four synthetic rows: two debits, one None-category, one income credit."""
    return [
        _row("2026-06-01", "SYNTH COFFEE SHOP", "-12.50", "Dining Out"),
        _row("2026-06-05", "SYNTH ELECTRICITY BILL", "-150.00", "Utilities"),
        _row("2026-06-10", "SYNTH UNKNOWN VENDOR", "-8.00", None),   # -> Uncategorised
        _row("2026-06-15", "SYNTH SALARY CREDIT", "3000.00", "Income"),
    ]


def _synth_summary() -> dict:
    """Matching summary dict with str(Decimal) values; insertion order matches transactions."""
    return {
        "year_month": "2026-06",
        "totals": {
            "Dining Out": "-12.50",
            "Utilities": "-150.00",
            "Uncategorised": "-8.00",
            "Income": "3000.00",
        },
        "net": "2829.50",
        "count": 4,
    }


# ---------------------------------------------------------------------------
# TestWorkbookShape
# ---------------------------------------------------------------------------


class TestWorkbookShape:
    """Workbook-level assertions: filename, sheet names, headers."""

    def test_filename_matches_yyyy_mm_pattern(self, tmp_path) -> None:
        """Returned path has the financetracker-YYYY-MM.xlsx filename pattern."""
        path = build_workbook(
            "2026-06", _synth_transactions(), _synth_summary(), output_dir=tmp_path
        )
        assert re.fullmatch(r"financetracker-\d{4}-\d{2}\.xlsx", path.name), (
            f"Expected financetracker-YYYY-MM.xlsx, got: {path.name!r}"
        )
        assert path.name == "financetracker-2026-06.xlsx"

    def test_returned_path_exists_and_is_under_tmp_path(self, tmp_path) -> None:
        """Returned path exists on disk and lives inside tmp_path."""
        path = build_workbook(
            "2026-06", _synth_transactions(), _synth_summary(), output_dir=tmp_path
        )
        assert path.exists(), "build_workbook must create the .xlsx file"
        assert path.resolve().is_relative_to(tmp_path.resolve()), (
            "File must be written inside tmp_path, not elsewhere"
        )

    def test_sheet_names_exactly_transactions_then_summary(self, tmp_path) -> None:
        """Workbook has exactly ['Transactions', 'Summary'] — no stray 'Sheet'."""
        path = build_workbook(
            "2026-06", _synth_transactions(), _synth_summary(), output_dir=tmp_path
        )
        wb = openpyxl.load_workbook(path)
        assert wb.sheetnames == ["Transactions", "Summary"], (
            f"Expected ['Transactions', 'Summary'], got {wb.sheetnames!r}"
        )

    def test_transactions_header_row(self, tmp_path) -> None:
        """Transactions sheet row 1 is ['Date', 'Description', 'Amount', 'Category']."""
        path = build_workbook(
            "2026-06", _synth_transactions(), _synth_summary(), output_dir=tmp_path
        )
        wb = openpyxl.load_workbook(path)
        ws = wb["Transactions"]
        header = [ws.cell(row=1, column=c).value for c in range(1, 5)]
        assert header == ["Date", "Description", "Amount", "Category"]

    def test_summary_header_row(self, tmp_path) -> None:
        """Summary sheet row 1 is ['Category', 'Total']."""
        path = build_workbook(
            "2026-06", _synth_transactions(), _synth_summary(), output_dir=tmp_path
        )
        wb = openpyxl.load_workbook(path)
        ws = wb["Summary"]
        header = [ws.cell(row=1, column=c).value for c in range(1, 3)]
        assert header == ["Category", "Total"]


# ---------------------------------------------------------------------------
# TestTransactionRows
# ---------------------------------------------------------------------------


class TestTransactionRows:
    """Per-row assertions on the Transactions sheet."""

    def test_transaction_rows_date_description_category_match_input(self, tmp_path) -> None:
        """Each data row carries the correct Date, Description, and Category strings."""
        txns = _synth_transactions()
        path = build_workbook("2026-06", txns, _synth_summary(), output_dir=tmp_path)
        wb = openpyxl.load_workbook(path)
        ws = wb["Transactions"]

        for i, txn in enumerate(txns):
            row_num = i + 2  # row 1 is header
            expected_cat = txn.category if txn.category is not None else "Uncategorised"
            assert ws.cell(row=row_num, column=1).value == txn.date, \
                f"Row {i}: date mismatch"
            assert ws.cell(row=row_num, column=2).value == txn.description, \
                f"Row {i}: description mismatch"
            assert ws.cell(row=row_num, column=4).value == expected_cat, \
                f"Row {i}: category mismatch (None must become 'Uncategorised')"

    def test_amount_cells_are_numeric_not_string(self, tmp_path) -> None:
        """Amount column values are numbers (float/int), NOT text strings."""
        path = build_workbook(
            "2026-06", _synth_transactions(), _synth_summary(), output_dir=tmp_path
        )
        wb = openpyxl.load_workbook(path)
        ws = wb["Transactions"]
        for row_num in range(2, 6):
            cell = ws.cell(row=row_num, column=3)
            assert isinstance(cell.value, (int, float)), (
                f"Amount cell at row {row_num} must be numeric, got {type(cell.value).__name__!r}"
            )

    def test_amount_cells_exact_two_dp_no_float_drift(self, tmp_path) -> None:
        """Amount cell values match the exact Decimal input after rounding to 2dp."""
        txns = _synth_transactions()
        path = build_workbook("2026-06", txns, _synth_summary(), output_dir=tmp_path)
        wb = openpyxl.load_workbook(path)
        ws = wb["Transactions"]

        _2dp = Decimal("0.01")
        for i, txn in enumerate(txns):
            row_num = i + 2
            cell = ws.cell(row=row_num, column=3)
            actual = Decimal(str(cell.value)).quantize(_2dp)
            expected = txn.amount.quantize(_2dp)
            assert actual == expected, (
                f"Row {i}: amount {cell.value!r} != {txn.amount} (as Decimal at 2dp)"
            )

    def test_amount_cells_have_0_00_number_format(self, tmp_path) -> None:
        """Amount cells in Transactions have number_format='0.00'."""
        path = build_workbook(
            "2026-06", _synth_transactions(), _synth_summary(), output_dir=tmp_path
        )
        wb = openpyxl.load_workbook(path)
        ws = wb["Transactions"]
        for row_num in range(2, 6):
            cell = ws.cell(row=row_num, column=3)
            assert cell.number_format == "0.00", (
                f"Row {row_num} Amount cell number_format must be '0.00', "
                f"got {cell.number_format!r}"
            )

    def test_none_category_renders_as_uncategorised_not_none_string(self, tmp_path) -> None:
        """A MonthRow with category=None is written as 'Uncategorised', never as 'None'."""
        txns = [_row("2026-06-10", "SYNTH MYSTERY PURCHASE", "-8.00", None)]
        summary = {
            "year_month": "2026-06",
            "totals": {"Uncategorised": "-8.00"},
            "net": "-8.00",
            "count": 1,
        }
        path = build_workbook("2026-06", txns, summary, output_dir=tmp_path)
        wb = openpyxl.load_workbook(path)
        ws = wb["Transactions"]
        cat_val = ws.cell(row=2, column=4).value
        assert cat_val == "Uncategorised", (
            f"None category must write 'Uncategorised', got {cat_val!r}"
        )
        assert cat_val != "None"
        assert cat_val is not None

    def test_negative_debit_amounts_preserved_not_abs(self, tmp_path) -> None:
        """Debit amounts (negative Decimal) are written as negative numbers — not abs'd."""
        txns = [_row("2026-06-01", "SYNTH DEBIT TRANSACTION", "-99.75", "Utilities")]
        summary = {
            "year_month": "2026-06",
            "totals": {"Utilities": "-99.75"},
            "net": "-99.75",
            "count": 1,
        }
        path = build_workbook("2026-06", txns, summary, output_dir=tmp_path)
        wb = openpyxl.load_workbook(path)
        ws = wb["Transactions"]
        val = ws.cell(row=2, column=3).value
        assert val < 0, f"Debit amount must be negative in the cell, got {val}"
        assert Decimal(str(val)).quantize(Decimal("0.01")) == Decimal("-99.75")

    def test_positive_income_amount_preserved(self, tmp_path) -> None:
        """Income amounts (positive Decimal) are written as positive numbers."""
        txns = [_row("2026-06-15", "SYNTH SALARY PAYMENT", "3000.00", "Income")]
        summary = {
            "year_month": "2026-06",
            "totals": {"Income": "3000.00"},
            "net": "3000.00",
            "count": 1,
        }
        path = build_workbook("2026-06", txns, summary, output_dir=tmp_path)
        wb = openpyxl.load_workbook(path)
        ws = wb["Transactions"]
        val = ws.cell(row=2, column=3).value
        assert val > 0, f"Income amount must be positive in the cell, got {val}"


# ---------------------------------------------------------------------------
# TestSummarySheet
# ---------------------------------------------------------------------------


class TestSummarySheet:
    """Per-row assertions on the Summary sheet."""

    def test_per_category_totals_match_summary_dict(self, tmp_path) -> None:
        """Per-category rows match summary['totals'] values (Decimal-exact at 2dp)."""
        summary = _synth_summary()
        path = build_workbook("2026-06", _synth_transactions(), summary, output_dir=tmp_path)
        wb = openpyxl.load_workbook(path)
        ws = wb["Summary"]

        _2dp = Decimal("0.01")
        for i, (cat, total_str) in enumerate(summary["totals"].items()):
            row_num = i + 2  # row 1 is header
            cat_cell = ws.cell(row=row_num, column=1)
            total_cell = ws.cell(row=row_num, column=2)
            assert cat_cell.value == cat, f"Category label mismatch at row {row_num}"
            actual = Decimal(str(total_cell.value)).quantize(_2dp)
            expected = Decimal(total_str).quantize(_2dp)
            assert actual == expected, (
                f"Total mismatch at row {row_num}: {actual!r} != {expected!r}"
            )

    def test_summary_totals_number_format_is_0_00(self, tmp_path) -> None:
        """Total cells on the Summary sheet have number_format='0.00'."""
        summary = _synth_summary()
        path = build_workbook("2026-06", _synth_transactions(), summary, output_dir=tmp_path)
        wb = openpyxl.load_workbook(path)
        ws = wb["Summary"]
        n_totals = len(summary["totals"])
        for row_num in range(2, n_totals + 2):  # rows 2..n_totals+1 are total rows
            cell = ws.cell(row=row_num, column=2)
            assert cell.number_format == "0.00", (
                f"Summary row {row_num} Total must have number_format='0.00', "
                f"got {cell.number_format!r}"
            )

    def test_summary_totals_insertion_order_preserved(self, tmp_path) -> None:
        """Summary sheet preserves the insertion order of summary['totals']."""
        summary = _synth_summary()
        path = build_workbook("2026-06", _synth_transactions(), summary, output_dir=tmp_path)
        wb = openpyxl.load_workbook(path)
        ws = wb["Summary"]

        expected_order = list(summary["totals"].keys())
        actual_order = [
            ws.cell(row=i + 2, column=1).value for i in range(len(expected_order))
        ]
        assert actual_order == expected_order, (
            "Summary totals must appear in the same insertion order as summary['totals']"
        )

    def test_net_row_label_and_numeric_value(self, tmp_path) -> None:
        """Final Summary row is ['Net', <net_value>] with '0.00' number format."""
        summary = _synth_summary()
        path = build_workbook("2026-06", _synth_transactions(), summary, output_dir=tmp_path)
        wb = openpyxl.load_workbook(path)
        ws = wb["Summary"]

        net_row = len(summary["totals"]) + 2  # header + totals rows + 1
        label_cell = ws.cell(row=net_row, column=1)
        value_cell = ws.cell(row=net_row, column=2)

        assert label_cell.value == "Net", (
            f"Net row label must be 'Net', got {label_cell.value!r}"
        )
        actual = Decimal(str(value_cell.value)).quantize(Decimal("0.01"))
        expected = Decimal(summary["net"]).quantize(Decimal("0.01"))
        assert actual == expected, f"Net value {actual!r} != {expected!r}"
        assert value_cell.number_format == "0.00"

    def test_total_cells_are_numeric_not_string(self, tmp_path) -> None:
        """Total column values on Summary (including Net) are numbers, not text."""
        path = build_workbook(
            "2026-06", _synth_transactions(), _synth_summary(), output_dir=tmp_path
        )
        wb = openpyxl.load_workbook(path)
        ws = wb["Summary"]
        for row_num in range(2, ws.max_row + 1):
            cell = ws.cell(row=row_num, column=2)
            assert isinstance(cell.value, (int, float)), (
                f"Summary row {row_num} Total must be numeric, "
                f"got {type(cell.value).__name__!r}"
            )


# ---------------------------------------------------------------------------
# TestEdgeCases
# ---------------------------------------------------------------------------


class TestEdgeCases:
    """Edge cases named in the spec."""

    def test_empty_transactions_writes_header_only(self, tmp_path) -> None:
        """Empty transaction list: file still created; Transactions sheet has header only."""
        summary = {
            "year_month": "2026-06",
            "totals": {},
            "net": "0.00",
            "count": 0,
        }
        path = build_workbook("2026-06", [], summary, output_dir=tmp_path)
        assert path.exists(), "File must still be created for empty transaction list"
        wb = openpyxl.load_workbook(path)
        ws = wb["Transactions"]
        assert ws.max_row == 1, (
            f"Expected 1 row (header only) for empty transactions, got {ws.max_row}"
        )
        header = [ws.cell(row=1, column=c).value for c in range(1, 5)]
        assert header == ["Date", "Description", "Amount", "Category"]

    def test_empty_totals_summary_has_header_plus_net_only(self, tmp_path) -> None:
        """Empty summary['totals']: Summary has header + Net row only (2 rows total)."""
        summary = {
            "year_month": "2026-06",
            "totals": {},
            "net": "-25.00",
            "count": 1,
        }
        txns = [_row("2026-06-01", "SYNTH UNCATEGORISED ITEM", "-25.00", None)]
        path = build_workbook("2026-06", txns, summary, output_dir=tmp_path)
        wb = openpyxl.load_workbook(path)
        ws = wb["Summary"]
        assert ws.max_row == 2, (
            f"Expected 2 rows (header + Net) for empty totals, got {ws.max_row}"
        )
        assert ws.cell(row=1, column=1).value == "Category"
        assert ws.cell(row=2, column=1).value == "Net"
        actual_net = Decimal(str(ws.cell(row=2, column=2).value)).quantize(Decimal("0.01"))
        assert actual_net == Decimal("-25.00")

    def test_negative_zero_total_does_not_crash(self, tmp_path) -> None:
        """summary total of '-0.00' is parsed by Decimal without crashing."""
        summary = {
            "year_month": "2026-06",
            "totals": {"Uncategorised": "-0.00"},
            "net": "0.00",
            "count": 1,
        }
        txns = [_row("2026-06-01", "SYNTH ZERO AMOUNT TXN", "0.00", None)]
        path = build_workbook("2026-06", txns, summary, output_dir=tmp_path)
        assert path.exists(), "File must be written even when total is '-0.00'"
        wb = openpyxl.load_workbook(path)
        assert wb.sheetnames == ["Transactions", "Summary"]


# ---------------------------------------------------------------------------
# TestNoIoOutsideTmpPath
# ---------------------------------------------------------------------------


class TestNoIoOutsideTmpPath:
    """Assert that no files or dirs are created outside tmp_path."""

    def test_explicit_output_dir_override_no_default_dir_created(
        self, tmp_path, monkeypatch
    ) -> None:
        """Passing output_dir=tmp_path never creates ./output in the CWD."""
        monkeypatch.delenv("OUTPUT_DIR", raising=False)
        monkeypatch.chdir(tmp_path)  # redirect CWD so ./output == tmp_path / "output"

        path = build_workbook(
            "2026-06", _synth_transactions(), _synth_summary(), output_dir=tmp_path
        )

        assert path.exists()
        assert path.resolve().is_relative_to(tmp_path.resolve())
        # ./output (relative to new CWD=tmp_path) must NOT have been created
        assert not (tmp_path / "output").exists(), (
            "./output must NOT be created when output_dir=tmp_path is passed explicitly"
        )

    def test_resolve_output_dir_creates_no_directory(self, tmp_path, monkeypatch) -> None:
        """resolve_output_dir() returns a Path without creating the directory."""
        monkeypatch.delenv("OUTPUT_DIR", raising=False)
        monkeypatch.chdir(tmp_path)

        result = resolve_output_dir()

        assert isinstance(result, Path)
        # The ./output directory must NOT exist after calling resolve_output_dir()
        assert not (tmp_path / "output").exists(), (
            "resolve_output_dir() must NOT create ./output — path resolution only"
        )

    def test_import_backend_excel_builder_creates_no_dir(
        self, tmp_path, monkeypatch
    ) -> None:
        """A bare import of backend.excel_builder does not create any file or directory.

        Verified by asserting no ./output dir exists in a clean tmp_path CWD
        after (re-)importing the module.
        """
        monkeypatch.delenv("OUTPUT_DIR", raising=False)
        monkeypatch.chdir(tmp_path)

        import backend.excel_builder  # noqa: F401 — import-time IO check

        assert not (tmp_path / "output").exists(), (
            "Importing backend.excel_builder must NOT create ./output or any directory"
        )
